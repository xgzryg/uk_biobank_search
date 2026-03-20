"""
UK Biobank Search - 从UK Biobank Showcase检索文献并导出Excel
"""

import asyncio
import re
import sys
from html import unescape
from pathlib import Path
from playwright.async_api import async_playwright
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


async def search_uk_biobank(topic: str, output_path: str) -> dict:
    """
    Search UK Biobank for a topic and export to Excel

    Args:
        topic: Search term (e.g., "diabetes", "cancer", "cardiovascular")
        output_path: Output Excel file path

    Returns:
        Dictionary with publications and applications counts
    """
    print(f"Searching UK Biobank for: {topic}")
    url = "https://biobank.ndph.ox.ac.uk/showcase/search.cgi"

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        page = await browser.new_page()

        try:
            # Navigate to search page
            await page.goto(url, timeout=30000)
            print("Page loaded")

            # Fill in search term
            await page.fill('input[name="searchterm"]', topic)
            print(f"Entered search term: {topic}")

            # Click search button
            await page.click('input[type="submit"][value="Search"]')
            print("Clicked search button")

            # Wait for results to load
            await page.wait_for_load_state("networkidle", timeout=30000)
            print("Results loaded")

            # Get page content
            html_content = await page.content()
            print("Retrieved page content")

        except Exception as e:
            print(f"Error during search: {e}")
            raise
        finally:
            await browser.close()

    # Decode HTML entities
    html_content = unescape(html_content)
    soup = BeautifulSoup(html_content, 'html.parser')

    # Parse publications
    print("Parsing publications...")
    publications = []
    pub_rows = soup.find_all('tr', id=re.compile(r'^p\d+$'))

    for row in pub_rows:
        cells = row.find_all('td')
        if len(cells) >= 5:
            pub_id = cells[0].get_text(strip=True)
            title = cells[1].get_text(strip=True)
            authors = cells[2].get_text(strip=True)
            year = cells[3].get_text(strip=True)
            journal = cells[4].get_text(strip=True)
            publications.append([pub_id, title, authors, year, journal])

    print(f"Found {len(publications)} publications")

    # Parse applications
    print("Parsing applications...")
    applications = []
    app_rows = soup.find_all('tr', id=re.compile(r'^a\d+$'))

    for row in app_rows:
        cells = row.find_all('td')
        if len(cells) >= 2:
            app_id = cells[0].get_text(strip=True)
            title = cells[1].get_text(strip=True)
            applications.append([app_id, title])

    print(f"Found {len(applications)} applications")

    # Create Excel workbook
    print("Creating Excel workbook...")
    wb = Workbook()

    # Header styling
    header_font_white = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    # Publications sheet
    ws_pub = wb.active
    ws_pub.title = "Publications"
    ws_pub.append(["Pub ID", "Title", "Authors", "Year", "Journal"])
    for cell in ws_pub[1]:
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for pub in publications:
        ws_pub.append(pub)
    ws_pub.column_dimensions['A'].width = 10
    ws_pub.column_dimensions['B'].width = 80
    ws_pub.column_dimensions['C'].width = 40
    ws_pub.column_dimensions['D'].width = 8
    ws_pub.column_dimensions['E'].width = 40

    # Applications sheet
    ws_app = wb.create_sheet("Applications")
    ws_app.append(["Application ID", "Title"])
    for cell in ws_app[1]:
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for app in applications:
        ws_app.append(app)
    ws_app.column_dimensions['A'].width = 15
    ws_app.column_dimensions['B'].width = 100

    # Save workbook
    wb.save(output_path)
    print(f"Excel file saved to: {output_path}")

    return {
        "topic": topic,
        "publications_count": len(publications),
        "applications_count": len(applications),
        "output_file": output_path
    }


def main():
    """Command line interface"""
    if len(sys.argv) < 3:
        print("=" * 50)
        print("UK Biobank Search - 文献检索与Excel导出")
        print("=" * 50)
        print("\n用法: python uk_biobank_search.py <搜索主题> <输出文件路径>")
        print("\n示例:")
        print('  python uk_biobank_search.py "diabetes" "C:\\Users\\Lenovo\\Desktop\\diabetes.xlsx"')
        print('  python uk_biobank_search.py "cancer" "C:\\Users\\Lenovo\\Desktop\\cancer.xlsx"')
        print('  python uk_biobank_search.py "cardiovascular" "C:\\Users\\Lenovo\\Desktop\\cardio.xlsx"')
        print("\n或直接运行:")
        print('  python -c "import asyncio; from uk_biobank_search import search_uk_biobank; '
              'asyncio.run(search_uk_biobank(\'diabetes\', \'output.xlsx\'))"')
        sys.exit(1)

    topic = sys.argv[1]
    output_path = sys.argv[2]

    result = asyncio.run(search_uk_biobank(topic, output_path))

    print("\n" + "=" * 50)
    print("检索完成!")
    print("=" * 50)
    print(f"搜索主题: {result['topic']}")
    print(f"文献数量: {result['publications_count']}")
    print(f"申请数量: {result['applications_count']}")
    print(f"输出文件: {result['output_file']}")


if __name__ == "__main__":
    main()
